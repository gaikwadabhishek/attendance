import time
import openpyxl as opx
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter,column_index_from_string
import os

def mainWork(absent,subject,cl_division):
	direc = os.path.dirname(__file__)
	print(direc)
	absent = list(map(int,absent.split(' ')))
	absent = list(set(absent))
	lec_times = ('7:50','8:50','10:00','11:00','13:00','14:00','15:10')

	strength = 80 # Get from database
	subject = subject.lower()
	lec_number = 3 # Needs better way to get

	book = opx.load_workbook(direc+'/media/'+cl_division+'.xlsx')
	sheet = book.get_sheet_by_name(subject)

	curr_col = 1
	while sheet[get_column_letter(curr_col)+'1'].value !=None:
		print('found!')
		curr_col += 1

	curr_col = get_column_letter(curr_col)

#	for i in range(1,strength+1):
#		sheet[next_col+str(i)].value = sheet[prev_col+str(i)].value
#		sheet[prev_col+str(i)].value = None

	sheet[curr_col+'1'].value = str(time.localtime().tm_mday) + '-' + str(time.localtime().tm_mon) + '-' + str(time.localtime().tm_year)

	redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
	greenFill = PatternFill(start_color='FF00FF00',end_color='FF00FF00',fill_type='solid')

	for i in range(1,strength+1):
		if i in absent:
			sheet[curr_col+str(i+1)].value = 'A'
			sheet[curr_col+str(i+1)].fill = redFill

		else:
			sheet[curr_col+str(i+1)].value = 'P'
			sheet[curr_col+str(i+1)].fill = greenFill
		print('Marked!',str(i))

	sheet[curr_col+str(strength+3)].value = '%.2f' % (((strength-len(absent))/strength) * 100)
	sheet[curr_col+str(strength+3)].value += '%'

	book.save(direc+'/media/'+cl_division+'.xlsx')
