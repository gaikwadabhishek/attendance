import time
import openpyxl as opx
from openpyxl.utils import get_column_letter

def mainWork(absent):
	absent = list(map(int,absent.split(' ')))
	lec_times = ('7:50','8:50','10:00','11:00','13:00','14:00','15:10')

	strength = 60 # Get from database
	subject = 'sdl' # Get subject code from database
	cl_division = 'TEA' # Get class and division from database
	lec_number = 3 # Needs better way to get

	book = opx.load_workbook(cl_division+'.xlsx')
	sheet = book.get_sheet_by_name(subject)

	curr_col = 1
	while sheet[get_column_letter(curr_col)+'1'].value !=None:
		curr_col += 1

	curr_col = get_column_letter(curr_col)

	sheet[curr_col+'1'].value = str(time.localtime().tm_mday) + '-' + str(time.localtime().tm_mon) + '-' + str(time.localtime().tm_year)

	for i in range(1,strength+1):
		if i in absent:
			sheet[curr_col+str(i+1)].value = 'A'
		else:
			sheet[curr_col+str(i+1)].value = 'P'
		print('Marked!',str(i))

	sheet[curr_col+str(strength+3)].value = '%.2f' % (((strength-len(absent))/strength) * 100)
	sheet[curr_col+str(strength+3)].value += '%'

	book.save(cl_division+'.xlsx')
